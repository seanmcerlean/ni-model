"""Build the normalized NISRA projection data and current model YAML.

Usage:
    python scripts/build_current_model.py /path/to/NPP24_ppp_coc.xlsx

The workbook is the official NISRA 2024-based principal projection summary.
Observed 2022-2024 components come from the normalized mid-year estimates
already checked into ``data/``.
"""

import argparse
import csv
import io
import zipfile
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parents[1]
OBSERVED_PATH = ROOT / "data" / "ni_population_components_2002_2024.csv"
PROJECTION_PATH = ROOT / "data" / "ni_population_projection_2024_2074.csv"
LGD_POPULATION_PATH = ROOT / "data" / "ni_census_2021_lgd_population.csv"
INTERNAL_MIGRATION_PATH = ROOT / "data" / "ni_internal_migration_lgd_2021.csv"
LGD_PROJECTION_PATH = ROOT / "data" / "ni_lgd_population_projection_2022_2047.csv"
COMMUNITY_BACKGROUND_PATH = (
    ROOT / "data" / "ni_census_2021_lgd_community_background.csv"
)
COMMUNITY_MIGRATION_PATH = (
    ROOT / "data" / "ni_internal_migration_lgd_2021_by_religion.csv"
)
MODEL_PATH = ROOT / "models" / "ni_current.yaml"

MIGRATION_RELIGION_MAP = {
    "Catholic": "catholic",
    "Protestant and Other Christian": "protestant",
    "Other Religions": "other",
    "No Religion": "none",
    "Religion Not Stated": "none",
}

INTEGRATION_RATES = [
    (0.35, "CATHOLIC", "NONE"),
    (0.65, "PROTESTANT", "NONE"),
    (0.03, "CATHOLIC", "PROTESTANT"),
    (0.03, "PROTESTANT", "CATHOLIC"),
    (3.2, "NONE", "CATHOLIC"),
    (2.8, "NONE", "PROTESTANT"),
    (0.2, "NONE", "OTHER"),
    (4.0, "OTHER", "CATHOLIC"),
    (3.5, "OTHER", "PROTESTANT"),
    (1.0, "OTHER", "NONE"),
]

MORTALITY_AGE_RATES = [
    (0, 0, 4.424556),
    (1, 4, 0.091557),
    (5, 9, 0.073626),
    (10, 14, 0.030976),
    (15, 24, 0.466414),
    (25, 34, 0.83605),
    (35, 44, 1.455992),
    (45, 54, 2.839962),
    (55, 64, 7.009875),
    (65, 74, 16.609164),
    (75, 84, 44.028801),
    (85, 130, 149.379433),
]


def _integer(value):
    return int(round(float(value)))


def _year(value):
    return int(str(value)[:4])


def extract_projection(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = workbook["PERSONS"]
    year_cells = next(
        sheet.iter_rows(min_row=7, max_row=7, min_col=2, max_col=52, values_only=True)
    )
    years = [_year(value) for value in year_cells]
    rows = {
        row[0]: row[1:]
        for row in sheet.iter_rows(
            min_row=8, max_row=20, min_col=1, max_col=52, values_only=True
        )
    }

    projection = []
    for index, year in enumerate(years):
        if year == 2024:
            continue
        row = {
            "year": year,
            "population_start": _integer(rows["Population at start"][index]),
            "births": _integer(rows["Births"][index]),
            "deaths": _integer(rows["Deaths"][index]),
            "immigration": _integer(rows["International migration inflows"][index])
            + _integer(rows["Cross border migration inflows [note 3]"][index]),
            "emigration": _integer(rows["International migration outflows"][index])
            + _integer(rows["Cross border migration outflows [note 3]"][index]),
            "net_migration": _integer(rows["Net migration"][index]),
            "population_end": _integer(rows["Population at end"][index]),
        }
        row["reconciliation_adjustment"] = row["population_end"] - (
            row["population_start"]
            + row["births"]
            - row["deaths"]
            + row["net_migration"]
        )
        projection.append(row)
    return projection


def read_observed():
    with OBSERVED_PATH.open(encoding="utf-8", newline="") as source:
        rows = list(csv.DictReader(source))
    return [
        {
            "year": int(row["period"][-4:]),
            "population_start": int(row["population_start"]),
            "births": int(row["births"]),
            "deaths": int(row["deaths"]),
            "immigration": "",
            "emigration": "",
            "net_migration": int(row["net_migration"]),
            "reconciliation_adjustment": int(row["other_changes"]),
            "population_end": int(row["population_end"]),
        }
        for row in rows
        if int(row["period"][-4:]) >= 2022
    ]


def read_normalized_projection():
    with PROJECTION_PATH.open(encoding="utf-8", newline="") as source:
        rows = list(csv.DictReader(source))
    return [
        {key: int(value) if value else "" for key, value in row.items()} for row in rows
    ]


def read_community_internal_migration():
    with COMMUNITY_MIGRATION_PATH.open(encoding="utf-8", newline="") as source:
        rows = list(csv.DictReader(source))
    return [
        {
            **row,
            "count": int(row["count"]),
            "source_population": int(row["source_population"]),
        }
        for row in rows
    ]


def write_projection(rows):
    fields = [
        "year",
        "population_start",
        "births",
        "deaths",
        "immigration",
        "emigration",
        "net_migration",
        "reconciliation_adjustment",
        "population_end",
    ]
    with PROJECTION_PATH.open("w", encoding="utf-8", newline="") as target:
        writer = csv.DictWriter(target, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def extract_lgd_projection(workbook_path):
    """Normalize the official sub-national population trajectory."""
    workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    rows = workbook["Flat"].iter_rows(values_only=True)
    headings = next(rows)
    records = [dict(zip(headings, row)) for row in rows]
    return [
        {
            "year": int(record["Year"].split("/")[1]),
            "code": record["Area_Code"],
            "location": record["Area_Name"],
            "population": int(record["MYE"]),
        }
        for record in records
        if record["Area_Code"].startswith("N09")
        and record["Category"] == "Population at End"
    ]


def write_lgd_projection(rows):
    with LGD_PROJECTION_PATH.open("w", encoding="utf-8", newline="") as target:
        writer = csv.DictWriter(
            target,
            fieldnames=("year", "code", "location", "population"),
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows(rows)


def read_lgd_population_targets():
    with LGD_POPULATION_PATH.open(encoding="utf-8", newline="") as source:
        locations = {
            row["code"]: row["location"].upper() for row in csv.DictReader(source)
        }
    with LGD_PROJECTION_PATH.open(encoding="utf-8", newline="") as source:
        rows = list(csv.DictReader(source))
    targets = {}
    for row in rows:
        targets.setdefault(int(row["year"]), {})[locations[row["code"]]] = int(
            row["population"]
        )
    return [
        {
            "year": year,
            "populations": populations,
            "evidence": "nisra_2022_lgd_principal_projection",
        }
        for year, populations in sorted(targets.items())
    ]


def extract_internal_migration(archive_path):
    with LGD_POPULATION_PATH.open(encoding="utf-8", newline="") as source:
        populations = {row["code"]: row for row in csv.DictReader(source)}
    with zipfile.ZipFile(archive_path) as archive:
        binary_source = archive.open("ODMG01NI-UK-LGD.csv")
        source = io.TextIOWrapper(binary_source, encoding="utf-8-sig")
        rows = csv.DictReader(source)
        flows = [
            {
                "source_code": row["Migrant one year ago area code"],
                "source": populations[row["Migrant one year ago area code"]][
                    "location"
                ],
                "destination_code": row["Area of residence code"],
                "destination": populations[row["Area of residence code"]]["location"],
                "count": int(row["Count"]),
                "source_population": int(
                    populations[row["Migrant one year ago area code"]]["count"]
                ),
            }
            for row in rows
            if row["Migrant one year ago area code"] in populations
            and row["Area of residence code"] in populations
            and row["Migrant one year ago area code"] != row["Area of residence code"]
        ]
    return sorted(flows, key=lambda row: (row["source_code"], row["destination_code"]))


def write_internal_migration(rows):
    fields = [
        "source_code",
        "source",
        "destination_code",
        "destination",
        "count",
        "source_population",
        "rate_per_1000",
    ]
    with INTERNAL_MIGRATION_PATH.open("w", encoding="utf-8", newline="") as target:
        writer = csv.DictWriter(target, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {**row, "rate_per_1000": _rate(row["count"], row["source_population"])}
            )


def _controlled_counts(counts, target):
    """Scale disclosure-affected subgroup cells to the published all-person total."""
    current = sum(counts.values())
    if current == 0:
        return {background: 0 for background in MIGRATION_RELIGION_MAP.values()}
    raw = {background: count * target / current for background, count in counts.items()}
    controlled = {background: int(value) for background, value in raw.items()}
    remainder = target - sum(controlled.values())
    for background in sorted(raw, key=lambda item: (raw[item] % 1, item), reverse=True)[
        :remainder
    ]:
        controlled[background] += 1
    return controlled


def extract_community_internal_migration(archive_path, total_flows):
    """Extract OD flows by published current religion as a background proxy."""
    with COMMUNITY_BACKGROUND_PATH.open(encoding="utf-8", newline="") as source:
        background_rows = list(csv.DictReader(source))
    populations = {
        (row["code"], background): int(row[background])
        for row in background_rows
        for background in MIGRATION_RELIGION_MAP.values()
    }
    codes = {row["code"] for row in background_rows}
    observed = {}
    with zipfile.ZipFile(archive_path) as archive:
        binary_source = archive.open("ODMG20NI-UK-LGD.csv")
        source = io.TextIOWrapper(binary_source, encoding="utf-8-sig")
        for row in csv.DictReader(source):
            source_code = row["Migrant one year ago area code"]
            destination_code = row["Area of residence code"]
            if source_code not in codes or destination_code not in codes:
                continue
            if source_code == destination_code:
                continue
            key = (source_code, destination_code)
            if key not in observed:
                observed[key] = {
                    background: 0 for background in MIGRATION_RELIGION_MAP.values()
                }
            background = MIGRATION_RELIGION_MAP[
                row["RELIGION_BELONG_TO_DVO_AGG5_label"]
            ]
            observed[key][background] += int(row["Count"])

    flows = []
    for total_flow in total_flows:
        key = (total_flow["source_code"], total_flow["destination_code"])
        controlled = _controlled_counts(observed[key], total_flow["count"])
        for background, count in controlled.items():
            if count == 0:
                continue
            flows.append(
                {
                    "source_code": total_flow["source_code"],
                    "source": total_flow["source"],
                    "destination_code": total_flow["destination_code"],
                    "destination": total_flow["destination"],
                    "religious_background": background,
                    "count": count,
                    "source_population": populations[(key[0], background)],
                }
            )
    return flows


def write_community_internal_migration(rows):
    fields = [
        "source_code",
        "source",
        "destination_code",
        "destination",
        "religious_background",
        "count",
        "source_population",
        "rate_per_1000",
    ]
    with COMMUNITY_MIGRATION_PATH.open("w", encoding="utf-8", newline="") as target:
        writer = csv.DictWriter(target, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {**row, "rate_per_1000": _rate(row["count"], row["source_population"])}
            )


def _rate(count, population):
    return round(count * 1000 / population, 6)


def _carry_forward_projection_year(row, year=2075):
    """Extend the official series by one clearly-labelled estimated year."""
    population_start = row["population_end"]
    population_end = (
        population_start
        + row["births"]
        - row["deaths"]
        + row["net_migration"]
        + row["reconciliation_adjustment"]
    )
    return {
        **row,
        "year": year,
        "population_start": population_start,
        "population_end": population_end,
        "evidence": "estimated_2074_carry_forward",
    }


def build_internal_migration_rules(internal_flows):
    return [
        {
            "rate": _rate(flow["count"], flow["source_population"]),
            "year_min": 2022,
            "destination": flow["destination"].upper(),
            "filters": {
                "location": flow["source"].upper(),
                "religious_background": flow["religious_background"].upper(),
            },
            "evidence": "census_2021_origin_destination_by_religion",
        }
        for flow in internal_flows
    ]


def build_model(rows, internal_flows):
    rows = list(rows)
    if rows and rows[-1]["year"] < 2075:
        rows.append(_carry_forward_projection_year(rows[-1]))
    births = []
    deaths = []
    migration = []
    for row in rows:
        year = row["year"]
        population = row["population_start"]
        after_births = population + row["births"]
        before_migration = after_births - row["deaths"]
        source = row.get(
            "evidence", "observed" if year <= 2024 else "principal_projection"
        )
        births.append(
            {
                "rate": _rate(row["births"], population),
                "year_min": year,
                "year_max": year,
                "filters": {},
                "evidence": source,
            }
        )
        deaths.append(
            {
                # Deaths run after births in the simulation engine.
                "rate": _rate(row["deaths"], after_births),
                "year_min": year,
                "year_max": year,
                "filters": {},
                "evidence": source,
            }
        )
        if row["immigration"] == "":
            migration.append(
                {
                    "rate": _rate(row["net_migration"], before_migration),
                    "year_min": year,
                    "year_max": year,
                    "filters": {},
                    "evidence": "observed_net_only",
                }
            )
        else:
            migration.extend(
                [
                    {
                        "rate": _rate(row["immigration"], before_migration),
                        "year_min": year,
                        "year_max": year,
                        "filters": {},
                        "evidence": source,
                        "flow": "in",
                    },
                    {
                        # Outflows run after inflows, so use the then-current
                        # population as the denominator.
                        "rate": -_rate(
                            row["emigration"], before_migration + row["immigration"]
                        ),
                        "year_min": year,
                        "year_max": year,
                        "filters": {},
                        "evidence": source,
                        "flow": "out",
                    },
                ]
            )

    return {
        "name": "NI Current – NISRA 2024 principal projection",
        "baseline_profile": "current",
        "baseline_population": 1_903_175,
        "description": (
            "Observed NISRA components for 2022–2024 followed by the official "
            "2024-based principal projection for 2025–2074, with 2075 estimated "
            "by carrying forward the final projected component rates. The population "
            "baseline uses Census 2021 marginals; internal relocation follows "
            "the Census 2021 LGD origin–destination pattern balanced toward the "
            "official 2022-based LGD population trajectory. This is a projection "
            "scenario, not a forecast."
        ),
        "baseline_year": 2021,
        "data_through": 2024,
        "projection_version": (
            "NISRA/ONS 2024-based principal projection plus estimated 2075 tail"
        ),
        "default_start_year": 2021,
        "default_end_year": 2075,
        "rate_jitter": 0,
        "random_seed": 42,
        "integration_rates": [
            {
                "rate": rate,
                "year_min": 2022,
                "destination": destination,
                "filters": {
                    "religious_background": source,
                    "age_min": 18,
                    "age_max": 44,
                },
                "evidence": "estimated_identity_transition",
            }
            for rate, source, destination in INTEGRATION_RATES
        ],
        "mortality_age_rates": [
            {"age_min": age_min, "age_max": age_max, "rate": rate}
            for age_min, age_max, rate in MORTALITY_AGE_RATES
        ],
        "birth_rates": births,
        "death_rates": deaths,
        "migration_rates": migration,
        "internal_migration_rates": build_internal_migration_rules(internal_flows),
        "lgd_population_targets": read_lgd_population_targets(),
        "lgd_relocation_calibration": {
            "strength": 0.65,
            "post_projection_strength": 0.15,
            "fade_years": 15,
            "evidence": "nisra_2022_lgd_principal_projection",
        },
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path, nargs="?")
    parser.add_argument("origin_destination_zip", type=Path, nargs="?")
    parser.add_argument("origin_destination_religion_zip", type=Path, nargs="?")
    parser.add_argument("--lgd-projection-workbook", type=Path)
    parser.add_argument("--normalized", action="store_true")
    args = parser.parse_args()
    if args.lgd_projection_workbook:
        write_lgd_projection(extract_lgd_projection(args.lgd_projection_workbook))
    if args.normalized:
        rows = read_normalized_projection()
        community_internal_flows = read_community_internal_migration()
        with MODEL_PATH.open("w", encoding="utf-8") as target:
            yaml.safe_dump(
                build_model(rows, community_internal_flows),
                target,
                sort_keys=False,
                allow_unicode=True,
                width=88,
            )
        return
    if not all(
        (
            args.workbook,
            args.origin_destination_zip,
            args.origin_destination_religion_zip,
        )
    ):
        parser.error("source workbooks are required unless --normalized is used")
    rows = read_observed() + extract_projection(args.workbook)
    internal_flows = extract_internal_migration(args.origin_destination_zip)
    community_internal_flows = extract_community_internal_migration(
        args.origin_destination_religion_zip, internal_flows
    )
    write_projection(rows)
    write_internal_migration(internal_flows)
    write_community_internal_migration(community_internal_flows)
    with MODEL_PATH.open("w", encoding="utf-8") as target:
        yaml.safe_dump(
            build_model(rows, community_internal_flows),
            target,
            sort_keys=False,
            allow_unicode=True,
            width=88,
        )


if __name__ == "__main__":
    main()
